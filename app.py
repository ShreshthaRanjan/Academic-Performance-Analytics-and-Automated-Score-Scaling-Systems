from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

app = Flask(__name__)

UPLOAD_FOLDER = "static/uploads"
PROCESSED_FOLDER = "processed"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload_file():
    file = request.files["file"]

    if not file:
        return jsonify({"error": "No file uploaded"})

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    df = pd.read_excel(filepath)

    # Validate columns
    if df.shape[1] != 5:
        return jsonify({"error": "Excel must have exactly 5 columns"})

    df.columns = ["Serial No", "Roll No", "Name", "Acquired Marks", "Grade"]

    # Sort by Acquired Marks descending
    df = df.sort_values(by="Acquired Marks", ascending=False)

    # Statistics for scored marks
    max_marks = df["Acquired Marks"].max()
    min_marks = df["Acquired Marks"].min()
    mean_marks = round(df["Acquired Marks"].mean(),2)
    std_marks = round(df["Acquired Marks"].std(),2)

    # Minimum passing (lowest DD)
    dd_students = df[df["Grade"] == "DD"]
    min_passing = dd_students["Acquired Marks"].min() if not dd_students.empty else None

    # Scaling
    df["Final Scaled Score"] = (100 / max_marks) * df["Acquired Marks"]
    df["Final Scaled Score"] = df["Final Scaled Score"].round(2)

    # Reorder columns
    df = df[["Serial No", "Roll No", "Name", "Acquired Marks",
             "Final Scaled Score", "Grade"]]

    # Stats for scaled marks
    max_scaled = df["Final Scaled Score"].max()
    min_scaled = round(df["Final Scaled Score"].min(),2)
    mean_scaled = round(df["Final Scaled Score"].mean(),2)
    std_scaled = round(df["Final Scaled Score"].std(),2)

    # Grade distribution
    grade_counts = df["Grade"].value_counts().to_dict()

    # Save processed file
    processed_path = os.path.join(PROCESSED_FOLDER, "processed_marks.xlsx")
    with pd.ExcelWriter(processed_path, engine="openpyxl") as writer:
        # Sheet 1: Processed Data
        df.to_excel(writer, sheet_name="Processed Data", index=False)
        # Sheet 2: Statistics
        stats_data = {
            "Metric": ["Max", "Min", "Mean", "Std Dev", "Min Passing"],
            "Scored Marks": [
                max_marks,
                min_marks,
                mean_marks,
                std_marks,
                min_passing
            ],
            "Scaled Marks": [
                max_scaled,
                min_scaled,
                mean_scaled,
                std_scaled,
                (100 / max_marks) * min_passing if min_passing else None
            ]
        }
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)
        workbook = writer.book
        sheet = writer.sheets["Statistics"]
        # Grade distribution below stats table
        row_start = 8
        sheet.cell(row=row_start, column=1, value="Grade")
        sheet.cell(row=row_start, column=2, value="Count")
        for i, (grade, count) in enumerate(grade_counts.items(), start=row_start + 1):
            sheet.cell(row=i, column=1, value=grade)
            sheet.cell(row=i, column=2, value=count)
        # Create Bar Chart
        chart = BarChart()
        chart.title = "Grade Distribution"
        chart.x_axis.title = "Grades"
        chart.y_axis.title = "Number of Students"
        data = Reference(sheet, min_col=2, min_row=row_start,
                     max_row=row_start + len(grade_counts))
        categories = Reference(sheet, min_col=1, min_row=row_start + 1,
                           max_row=row_start + len(grade_counts))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, "E10")

    return jsonify({
    "scored_stats": {
        "max": float(max_marks),
        "min": float(min_marks),
        "mean": float(mean_marks),
        "std": float(std_marks),
        "min_passing": float(min_passing) if min_passing is not None else None
    },
    "scaled_stats": {
        "max": float(max_scaled),
        "min": float(min_scaled),
        "mean": float(mean_scaled),
        "std": float(std_scaled),
        "min_passing": float((100 / max_marks) * min_passing) if min_passing is not None else None
    },
    "grade_distribution": {k: int(v) for k, v in grade_counts.items()}
    })



@app.route("/download")
def download_file():
    return send_file("processed/processed_marks.xlsx", as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
